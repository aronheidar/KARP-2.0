# -*- coding: utf-8 -*-
"""
build_hms_2027.py — vinnur opinbera „Fasteignamat 2027"-pakka HMS (xlsx, handsótt af hms.is 19.8.2026,
hráskrár í gogn/hms/raw/) í þrjú létt JSON:

  web/public/gogn/hms/matssvaedi_2027.json   ← 197 matssvæði íbúðarhúsnæðis + 76 undirmatssvæði:
                                               heiti, sérbýlis-/fjölbýlisstuðull, % breyting 2027, meðalverð á m²
                                               + nákvæmni gildandi fasteignamats (fmat vs kaupverð) úr sölu-úrtakinu
  gogn/hms/matssvaedi_punktar.json           ← þynnt sölu-úrtak [x, y, matssvæði] í ISN93 (EPSG:3057) fyrir
                                               k-NN vörpun heimilisfangs → matssvæði í build_hnit.js (EKKI shippað)
  gogn/hms_fasteignamat_2027.json            ← samantektir: fjöldi + fasteignamat 2026/2027 per tegund og per
                                               sveitarfélag×tegund (fyrir /fasteignir/, SSG um @gogn)
  gogn/hms/einingar.json                     ← eignar-lýsing per FASTNUM úr sölu-úrtakinu (hæð, hæðir eignar, lyftuhús,
                                               svalir m², bílskúr m², baðkör, sturtur) fyrir ~42k einingar → build_fasteignaskra.js
                                               sameinar inn í fasteignaskra/<pn>.json (eignaspjald). Breytir EKKI matinu —
                                               mælt 19.8.2026: þessir þættir bera ±1% umfram svæði+stærð+aldur.
  web/public/gogn/hms/sumarhus_2027.json     ← sumarhús: 142 matssvæði (stuðull, % breyting, meðalverð m²) + ~2.780 sölur 2016→2026 + fmat-nákvæmni
                                               (svæði, kaupverð, m², byggár, lóð, við vatn, hitaveita) fyrir sumarhúsamat

Sölu-úrtakið (2027_gagnasafn_ibudir2.xlsx, 49.656 sölur 2021-01→2026-02, 81 dálkar) er matslíkans-gagnasafn
HMS — ekki eignagrunnur. ⚠ est/aest eru EKKI verðmöt í kr (32–41% frávik á 2026-sölum, óþekkt eining) og eru
ekki notuð. `hverfi` = matssvæðis-Nr (staðfest: 189 kóðar, allir í matssvæða-töflunni).

KEYRSLA: python skriptur/build_hms_2027.py   (openpyxl; ~40 s)   → svo: node skriptur/build_hnit.js --from-disk
"""
import os, json, glob, datetime, statistics, collections
import openpyxl
import sys
try: sys.stdout.reconfigure(encoding='utf-8')   # Windows-konsóll (cp1252) þolir ekki ⚠/→ í print
except Exception: pass

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, 'gogn', 'hms', 'raw')
OUT_WEB = os.path.join(ROOT, 'web', 'public', 'gogn', 'hms')
OUT_GOGN = os.path.join(ROOT, 'gogn')
os.makedirs(OUT_WEB, exist_ok=True); os.makedirs(os.path.join(ROOT, 'gogn', 'hms'), exist_ok=True)

def finna(mynstur):
    m = glob.glob(os.path.join(RAW, '*' + mynstur + '*.xlsx'))
    if not m: raise SystemExit('vantar hráskrá: ' + mynstur + ' í ' + RAW)
    return m[0]

def num(v):
    try: return None if v is None or v == '' else float(v)
    except Exception: return None

def s(v): return str(v).strip() if v is not None else ''

NU = datetime.datetime.now().isoformat(timespec='seconds')

# ── 1) Matssvæði + undirmatssvæði ─────────────────────────────────────────────
wb = openpyxl.load_workbook(finna('Matssv'), read_only=True)
svaedi = {}
for r in list(wb['Íbúðarhúsnæði'].iter_rows(values_only=True))[1:]:
    if r[0] is None: continue
    svaedi[s(r[0])] = {'heiti': s(r[1]), 'st_ser': num(r[2]), 'st_fjol': num(r[3]), 'br_ser': num(r[4]), 'br_fjol': num(r[5]), 'm2_ser': num(r[6]), 'm2_fjol': num(r[7])}
undir = {}
for r in list(wb['Íbúðarhúsnæði undirmatssv'].iter_rows(values_only=True))[1:]:
    if r[0] is None: continue
    undir[s(r[0])] = {'heiti': s(r[1]), 'st_ser': num(r[2]), 'st_fjol': num(r[3])}
print('matssvæði:', len(svaedi), '· undirmatssvæði:', len(undir))

# ── 2) Sölu-úrtakið: punktar per matssvæði (þynnt) + nákvæmni fmat ───────────
ws = openpyxl.load_workbook(finna('gagnasafn_ibudir'), read_only=True)['Sheet1']
it = ws.iter_rows(values_only=True); H = list(next(it)); ix = {h: i for i, h in enumerate(H)}
g = lambda r, k: r[ix[k]]
GRID = 100  # m — ein sala per 100 m reit per svæði dugar fyrir k-NN, þynnir 49.656 → ~fjórðung
sed = set(); punktar = []; einingar = {}
fmat_err = []; n_alls = 0; hverfi_ekki = collections.Counter()
FRA = datetime.datetime(2025, 3, 1)
for r in it:
    n_alls += 1
    x, y, hv = num(g(r, 'x')), num(g(r, 'y')), s(g(r, 'hverfi'))
    if x and y and hv in svaedi:
        key = (hv, int(x // GRID), int(y // GRID))
        if key not in sed:
            sed.add(key); punktar.append([round(x), round(y), int(hv)])
    elif hv: hverfi_ekki[hv] += 1
    # eignar-lýsing per fastanúmer (nýjasta sala vinnur — lykkjan er í tímaröð? treystum ekki: geymum með dagsetningu)
    fn = g(r, 'fastnum'); dd = g(r, 'utgdag')
    if fn and dd:
        prev = einingar.get(int(fn))
        if not prev or prev[0] < dd:
            # haednr = hæð eignarinnar · fjhaed = fjöldi hæða EIGNARINNAR (ekki hússins — 47,6% sala hafa haednr>fjhaed) ·
            # lyftuhus 0/1 (lyfta = FJÖLDI lyfta 0–5) · fjbkar/fjsturt = baðkör/sturtur (summan er EKKI fjöldi baðherbergja)
            einingar[int(fn)] = [dd, [int(num(g(r,'haednr')) or 0), int(num(g(r,'fjhaed')) or 0), 1 if (num(g(r,'lyftuhus')) or num(g(r,'lyfta'))) else 0, round(num(g(r,'svalm2')) or 0, 1), round(num(g(r,'bilskurm2')) or 0, 1), int(num(g(r,'fjbkar')) or 0), int(num(g(r,'fjsturt')) or 0)]]
    # nákvæmni gildandi fasteignamats: sölur sl. 12 mán fyrir matsdag, án nýbygginga
    d, kv, fm, ny = g(r, 'utgdag'), num(g(r, 'kaupverd')), num(g(r, 'fmat')), g(r, 'nybygging')
    if d and d >= FRA and kv and kv > 0 and fm and not ny:
        fmat_err.append(abs(fm / kv - 1))
fmat_err.sort()
q = lambda p: fmat_err[int(len(fmat_err) * p)]
nakv = {'n': len(fmat_err), 'timabil': '3/2025–2/2026', 'midgildi': round(statistics.median(fmat_err), 4), 'p75': round(q(0.75), 4),
        'innan10': round(sum(e <= 0.10 for e in fmat_err) / len(fmat_err), 4), 'innan20': round(sum(e <= 0.20 for e in fmat_err) / len(fmat_err), 4),
        'lysing': 'Gildandi fasteignamat HMS (fmat) borið saman við þinglýst kaupverð sömu eigna í sölu-úrtaki HMS fyrir fasteignamat 2027 — sölur 3/2025–2/2026 án nýbygginga. ⚠ Ekki strangt out-of-sample: fasteignamatið var kvarðað á hluta þessara sala.'}
print('sölur alls:', n_alls, '· þynntir punktar:', len(punktar), '· hverfi utan töflu:', dict(hverfi_ekki) or 'engin')
print('fmat vs kaupverð:', nakv)

with open(os.path.join(OUT_WEB, 'matssvaedi_2027.json'), 'w', encoding='utf-8') as f:
    json.dump({'updated': NU, 'heimild': 'HMS — Fasteignamat 2027: Matssvæði vefsjá, tölfræði (hms.is), handsótt 19.8.2026', 'svaedi': svaedi, 'undir': undir, 'fmat_nakvaemni': nakv}, f, ensure_ascii=False, separators=(',', ':'))
with open(os.path.join(ROOT, 'gogn', 'hms', 'einingar.json'), 'w', encoding='utf-8') as f:
    json.dump({'updated': NU, 'svid': ['haed', 'fjhaed_eignar', 'lyftuhus', 'svalir_m2', 'bilskur_m2', 'badkor', 'sturtur'], 'n': len(einingar), 'heimild': 'HMS — 2027_gagnasafn_ibudir2.xlsx (sölu-úrtak matslíkans), nýjasta sala per FASTNUM', 'e': {str(k): v[1] for k, v in einingar.items()}}, f, separators=(',', ':'))
print('einingar.json:', len(einingar), 'fastanúmer með eignar-lýsingu')
with open(os.path.join(ROOT, 'gogn', 'hms', 'matssvaedi_punktar.json'), 'w', encoding='utf-8') as f:
    json.dump({'updated': NU, 'crs': 'EPSG:3057 (ISN93 / Lambert 1993)', 'grid_m': GRID, 'n': len(punktar), 'heimild': 'HMS — 2027_gagnasafn_ibudir2.xlsx (sölu-úrtak matslíkans), x/y/hverfi', 'punktar': punktar}, f, separators=(',', ':'))

# ── 2b) SUMARHÚS: 142 matssvæði + 2.799 sölur (2016-01→2026-02) fyrir sumarhúsamat (web/src/lib/sumarhusamat.mjs) ──
shz = {}
for r in list(wb['Sumarhús'].iter_rows(values_only=True))[1:]:
    if r[0] is None: continue
    shz[s(r[0])] = {'heiti': s(r[1]), 'studull': num(r[2]), 'br': num(r[3]), 'm2': num(r[4])}
wsh = openpyxl.load_workbook(finna('sumarh'), read_only=True); wsh = wsh[wsh.sheetnames[0]]
sit = wsh.iter_rows(values_only=True); SH = list(next(sit)); six = {h: i for i, h in enumerate(SH)}
gs = lambda r, k: r[six[k]]
solur = []; shz_ekki = collections.Counter(); sh_fm = []   # sh_fm: |fasteignamat/kaupverð − 1| sl. ~2 ár (viðmið f. Karp-bakpróf)
for r in sit:
    if r[0] is None: continue
    d, kv, m2, hv = gs(r, 'utgdag'), num(gs(r, 'kaupverd')), num(gs(r, 'sumarhus_m2')), s(gs(r, 'hverfi'))
    if not (d and kv and kv > 0 and m2 and m2 > 8): continue
    if gs(r, 'nybygging') or gs(r, 'seldadhluta') or gs(r, 'milli_skyldra'): continue
    if hv not in shz: shz_ekki[hv] += 1; continue
    fm = num(gs(r, 'fmat'))
    if fm and fm > 0 and d.strftime('%Y-%m-%d') >= '2024-03-01': sh_fm.append(abs(fm / kv - 1))
    solur.append({'d': d.strftime('%Y-%m-%d'), 'hv': int(hv), 'kv': int(kv), 'm2': round(m2, 1), 'ar': int(num(gs(r, 'byggar')) or 0) or None,
                  'lod': int(num(gs(r, 'lodpflm')) or 0) or None, 'eign': 1 if gs(r, 'eignarlod') else 0, 'vatn': 1 if gs(r, 'vid_vatn') else 0,
                  'hiti': 1 if gs(r, 'hitaveita') else 0, 'raf': 1 if gs(r, 'rafveita') else 0, 'ppm': round(kv * 1000 / m2)})
solur.sort(key=lambda x: x['d'], reverse=True)
sh_fm.sort()
fm_nakv = {'n': len(sh_fm), 'timabil': '2024-03→2026-02', 'midgildi': round(sh_fm[len(sh_fm) // 2], 3), 'p75': round(sh_fm[(len(sh_fm) * 3) // 4], 3),
           'innan20': round(sum(1 for e in sh_fm if e <= 0.2) / len(sh_fm), 3), 'innan30': round(sum(1 for e in sh_fm if e <= 0.3) / len(sh_fm), 3),
           'lysing': 'Opinbert fasteignamat sumarhúss (fmat í gagnasafninu) borið saman við kaupverð sömu sölu — |fmat/kaupverð − 1|. Viðmið fyrir bakpróf Karp-sumarhúsamatsins.'} if sh_fm else None
tb = (solur[-1]['d'][:7] + '→' + solur[0]['d'][:7]) if solur else ''
with open(os.path.join(OUT_WEB, 'sumarhus_2027.json'), 'w', encoding='utf-8') as f:
    json.dump({'updated': NU, 'heimild': 'HMS — Fasteignamat 2027: matssvæði sumarhúsa + gagnasafn sumarhúsamats (' + str(len(solur)) + ' nothæfar sölur ' + tb + '; nýbyggingar, hlutasölur og sölur milli skyldra síaðar burt). Kaupverð í þ.kr, ppm í kr/m².', 'svaedi': shz, 'fmat_nakvaemni': fm_nakv, 'solur': solur}, f, ensure_ascii=False, separators=(',', ':'))
print('sumarhús fmat vs kaupverð:', fm_nakv)
print('sumarhús: svæði', len(shz), '· sölur nothæfar', len(solur), '· svæði utan töflu:', dict(shz_ekki) or 'engin')

# ── 3) Samantektir: fjöldi + fasteignamat 2026/2027 ──────────────────────────
def lesa(mynstur, lyklar):
    wb = openpyxl.load_workbook(finna(mynstur), read_only=True); ws = wb[wb.sheetnames[0]]
    out = []; sidasta = ''
    for r in list(ws.iter_rows(values_only=True))[1:]:
        # sveitarfélags-dálkurinn er SAMFELLDUR (merged) í xlsx → nafnið stendur aðeins í fyrstu röð hvers hóps
        if lyklar[0] == 'sv':
            if r[0] is not None and s(r[0]): sidasta = s(r[0])
            if not sidasta or r[1] is None: continue
            r = (sidasta,) + tuple(r[1:])
        elif r[0] is None: continue
        out.append({k: (s(r[i]) if k in ('sv', 'teg', 'flokkur') else num(r[i])) for i, k in enumerate(lyklar)})
    return out
# Raðir án fjölda eru NEÐANMÁLSGREINAR í xlsx (t.d. Grindavíkur-fyrirvarinn) — haldið sem athugasemdum, ekki gögnum
athugasemdir = []
def hreinsa(rows, lykill):
    out = []
    for r in rows:
        hefurTolu = any(r.get(k) is not None for k in ('fjoldi', 'mat2026', 'mat2027'))
        if not hefurTolu and r.get(lykill) and len(r[lykill]) > 40: athugasemdir.append(r[lykill])
        elif hefurTolu: out.append(r)
    return out
byTeg = hreinsa(lesa('Eftirtegundeigna', ['teg', 'fjoldi', 'mat2026', 'mat2027', 'breyting']), 'teg')
bySvTeg = hreinsa(lesa('Sveitarf', ['sv', 'teg', 'fjoldi', 'mat2026', 'mat2027', 'breyting']), 'sv')
ibFlokkun = hreinsa(lesa('Íbúðareignireftirflokkun', ['flokkur', 'fjoldi', 'mat2026', 'mat2027', 'breyting']), 'flokkur')
atvFlokkun = hreinsa(lesa('Atvinnueignireftirflokkun', ['flokkur', 'fjoldi', 'mat2026', 'mat2027', 'breyting']), 'flokkur')
athugasemdir = list(dict.fromkeys(a.strip() for a in athugasemdir))
# „Samtals"-raðirnar í xlsx eru FORMÚLUR (=SUM…), ekki gildi → reiknum heildina sjálf úr tegundaröðunum
def samtala(rows):
    fj = sum(r['fjoldi'] or 0 for r in rows); m26 = sum(r['mat2026'] or 0 for r in rows); m27 = sum(r['mat2027'] or 0 for r in rows)
    return {'fjoldi': fj, 'mat2026': m26, 'mat2027': m27, 'breyting': round((m27 / m26 - 1) * 100, 1) if m26 else None}
byTeg = [r for r in byTeg if not r['teg'].startswith('Samtals')]
samtals = samtala(byTeg)
print('samtals (reiknað):', samtals)
print('athugasemdir úr xlsx:', athugasemdir)
with open(os.path.join(OUT_GOGN, 'hms_fasteignamat_2027.json'), 'w', encoding='utf-8') as f:
    json.dump({'updated': NU, 'heimild': 'HMS — Fasteignamat 2027, samantektir (hms.is), handsótt 19.8.2026. Mat í kr.', 'athugasemdir': athugasemdir, 'samtals': samtals, 'byTeg': byTeg, 'bySvTeg': bySvTeg, 'ibudirFlokkun': ibFlokkun, 'atvinnuFlokkun': atvFlokkun}, f, ensure_ascii=False, separators=(',', ':'))
print('samantektir: tegundir', len(byTeg), '· sveitarfélag×tegund', len(bySvTeg), '· íbúðaflokkun', len(ibFlokkun), '· atvinnuflokkun', len(atvFlokkun))
print('OK')
